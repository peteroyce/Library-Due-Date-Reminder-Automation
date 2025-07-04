"""
Standalone Python demo of the library due-date reminder logic.

This script mirrors what the UiPath workflow (Main.xaml) does:
  1. Read an Excel file with library records
  2. Check which books are due within N days
  3. Generate reminder emails for overdue/near-due books

Run:
  pip install openpyxl
  python demo/reminder_logic.py --file "Library Automation System.xlsx" --days 3
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)


EMAIL_TEMPLATE = """Subject: Library Due Date Reminder

Dear {student_name},

This is a reminder that the book "{book_name}" you borrowed is due on {due_date}.

{urgency_line}

Please return the book to avoid any penalties.

Best regards,
Library Management System
"""


def load_library_data(filepath: str) -> list[dict]:
    """Load library records from an Excel file."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        records.append(record)
    wb.close()
    return records


def check_due_dates(records: list[dict], threshold_days: int = 3) -> list[dict]:
    """Filter records that are due within the threshold or overdue."""
    today = datetime.now()
    alerts = []

    for record in records:
        due_date = record.get("Due Date")
        if due_date is None:
            continue

        if isinstance(due_date, str):
            try:
                due_date = datetime.strptime(due_date, "%d/%m/%Y")
            except ValueError:
                continue

        days_left = (due_date - today).days

        if days_left <= threshold_days:
            alerts.append({
                "student_name": record.get("Student Name", record.get("Patron name", "Unknown")),
                "book_name": record.get("Book Name", record.get("Book title", "Unknown")),
                "due_date": due_date.strftime("%d/%m/%Y"),
                "email": record.get("Email ID", record.get("Email", "")),
                "days_left": days_left,
                "is_overdue": days_left < 0,
            })

    return alerts


def generate_email(alert: dict) -> str:
    """Generate email content for a single alert."""
    if alert["is_overdue"]:
        urgency = f"This book is OVERDUE by {abs(alert['days_left'])} day(s). A penalty may apply."
    elif alert["days_left"] == 0:
        urgency = "This book is due TODAY. Please return it immediately."
    else:
        urgency = f"You have {alert['days_left']} day(s) remaining before the due date."

    return EMAIL_TEMPLATE.format(
        student_name=alert["student_name"],
        book_name=alert["book_name"],
        due_date=alert["due_date"],
        urgency_line=urgency,
    )


def main():
    parser = argparse.ArgumentParser(description="Library due-date reminder system")
    parser.add_argument("--file", required=True, help="Path to the Excel file")
    parser.add_argument("--days", type=int, default=3, help="Alert threshold in days (default: 3)")
    parser.add_argument("--dry-run", action="store_true", help="Print emails instead of sending")
    args = parser.parse_args()

    if not Path(args.file).exists():
        print(f"File not found: {args.file}")
        sys.exit(1)

    records = load_library_data(args.file)
    print(f"Loaded {len(records)} library records")

    alerts = check_due_dates(records, threshold_days=args.days)
    print(f"Found {len(alerts)} books due within {args.days} days")

    for alert in alerts:
        status = "OVERDUE" if alert["is_overdue"] else f"{alert['days_left']}d left"
        print(f"  [{status}] {alert['book_name']} — {alert['student_name']} ({alert['email']})")

        if args.dry_run:
            print(generate_email(alert))
            print("---")

    if not alerts:
        print("No reminders to send.")


if __name__ == "__main__":
    main()


def validate_0(data):
    """Validate: add data validation"""
    return data is not None
